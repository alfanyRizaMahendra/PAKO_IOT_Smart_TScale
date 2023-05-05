#!/usr/bin/python

from PyQt5 import QtCore, QtGui, QtWidgets, uic
from PyQt5.QtGui import *
from PyQt5.QtCore import *
from PyQt5.QtWidgets import *

#import RPi.GPIO as GPIO
import mysql.connector
#import serial
import random
import datetime
import os
import calendar
import sys
import threading
import numpy as np
import time

from openpyxl import Workbook
from openpyxl import load_workbook

now = datetime.datetime.now()
adjust_height = 40

weight_list = [2.6, 3.2, 59.96] # red box, green_box, Trolley
number_of_green = 5
number_of_red = 12

berat_timbangan = ["920.53", "780.26", "888.99", "876.88", "595.44", "788.66", "546.78", "984.57", "877.60", "766.99"]

def get_scale_value():
        #ser = serial.Serial('/dev/ttyUSB0',115200, timeout = 1)
        #ser.flush()
        #line = ser.readline().decode('utf-8').rstrip().replace(" ","").replace(",","").replace("STGS","").replace("kg","")
        # if line[:5]=="TG0.k":
        #         line = line[5:]
        # if line[:4] == "USGS" or line[:4] == "TSGS":
        #         line = line[4:]
        # if line[:3] == "SGS":
        #         line = line[3:]
        # if line[:3] == "SS":
        #         line = line[2:]
        # if line[:1] == "S" or line[:1] == "g":
        #         line = line[1:]
        # if line[-1:] == 'k' or line[-1:] == 'g':
        #         line = line[:-1]
        line = random.randint(0, 2000)
        print(line)
        return line
                
class meltingChip(QWidget):
        def db_connect(self,db):
                # DB settinwg 
                mydb = db.connect(
                        host = "localhost",
                        user = "root",
                        password = "root",
                        database = "tscale"
                        #connect_timeout=60
                )
                return mydb
        
        def __init__(self, timer):
                super().__init__()

                self.timer = timer
                self.separator_count = 0
                self.melting_count = 0
                self.db = mysql.connector
                self.mydb = self.db_connect(self.db)
                self.count_error = 0

                self.timer.timeout.connect(self.update_display)
                self.timer.start(50)

                #define variable
                self.convert_item_to_code = {}
                self.count = 0
                
                self.resize(1024, 600)
                _translate = QtCore.QCoreApplication.translate
                
                # PAKOAKUINA 2W Title
                self.pushButton = QtWidgets.QPushButton(self)
                self.pushButton.setGeometry(QtCore.QRect(-30, 0, 1054, 91))
                self.pushButton.setStyleSheet("*{background-color : #67ADE2;\n"
                                                "color : #000;\n"
                                                "font-weight : bold; \n"
                                                "font-size : 50px;\n"
                                                "text-align: right;\n"
                                                "padding-right : 50px}")
                self.pushButton.setObjectName("pushButton")

                # Separator button
                # self.pushButton_3 = QtWidgets.QPushButton(self, clicked = lambda:self.go_to_main_window2())
                # self.pushButton_3.setGeometry(QtCore.QRect(420, 100, 200, 60))
                # self.pushButton_3.setStyleSheet("background-color : #F0EB8D;\n"
                #                                 "color : #000;\n"
                #                                 "font-weight : bold; \n"
                #                                 "font-size : 30px;\n"
                #                                 "text-align: center;")
                # self.pushButton_3.setObjectName("pushButton_3")
                
                # Pako Logo 
                #self.label_image = QtWidgets.QLabel(self)
                #self.label_image.setGeometry(QtCore.QRect(-30, 0, 300, 60))
                #self.label_image.setPixmap(QPixmap('logo.png'))

                # choosing items
                self.label = QtWidgets.QLabel(self)
                self.label.setGeometry(QtCore.QRect(20, 220-adjust_height, 271, 91))
                self.label.setStyleSheet("font-size : 30px;\n")
                self.label.setObjectName("label")
                
                item = self.get_query("""SELECT * FROM master_data WHERE FLAG=1""")
                item2 = self.get_query("""SELECT * FROM master_data WHERE FLAG=2""")
                self.list_items_melt = []
                self.list_separator = []
                self.item_code_melt = []
                self.item_code_separator = []
                for i in range(len(item)):
                        self.list_items_melt.append(item[i][2])
                        self.item_code_melt.append(item[i][1])
                        self.convert_item_to_code.update({item[i][2]:item[i][1]})   

                for i in range(len(item2)):  
                        self.list_separator.append(item2[i][2])
                        self.item_code_separator.append(item2[i][1]) 
                        self.convert_item_to_code.update({item2[i][2]:item2[i][1]})   
                
                print(self.convert_item_to_code)
                self.comboBox = QtWidgets.QComboBox(self)
                self.comboBox.setGeometry(QtCore.QRect(20, 290-adjust_height, 365, 71))
                self.comboBox.setStyleSheet("font-size : 40px;\n""")
                self.comboBox.setObjectName("comboBox")
                self.comboBox.setCurrentText(_translate("MainWindow", "Chips NFE"))

                # Choosing Melting number
                self.label_4 = QtWidgets.QLabel(self)
                self.label_4.setGeometry(QtCore.QRect(450, 140-adjust_height, 271, 51))
                self.label_4.setStyleSheet("font-size : 30px;\n")
                self.label_4.setObjectName("label_4")
                
                list_melting_numbers=['melting 1','melting 2','melting 3','melting 4', "separator"]
                print(list_melting_numbers[0][:7])
                self.comboBox_3 = QtWidgets.QComboBox(self)
                self.comboBox_3.setGeometry(QtCore.QRect(450, 190-adjust_height, 280, 51))
                self.comboBox_3.setStyleSheet("font-size:40px;")
                self.comboBox_3.setObjectName("comboBox_3")
                self.comboBox_3.addItems(list_melting_numbers)

                # Choosing amount of red and green box 
                self.label_8 = QtWidgets.QLabel(self)
                self.label_8.setGeometry(QtCore.QRect(50, 370-adjust_height, 110, 35))
                self.label_8.setStyleSheet("font-size : 30px;\n")
                self.label_8.setObjectName("label_8")
                self.label_8.setVisible(False)

                self.label_7 = QtWidgets.QLabel(self)
                self.label_7.setGeometry(QtCore.QRect(230, 370-adjust_height, 110, 35))
                self.label_7.setStyleSheet("font-size : 30px;\n")
                self.label_7.setObjectName("label_7")
                self.label_7.setVisible(False)
                
                self.red_choice = []
                self.green_choice = []
                for self.i in range(1,number_of_red+1):
                        self.red_choice.append(str(self.i))
                        if(self.i < number_of_green+1):
                                self.green_choice.append(str(self.i))
                self.comboBox_4 = QtWidgets.QComboBox(self)
                self.comboBox_4.setGeometry(QtCore.QRect(20, 410-adjust_height, 150, 51))
                self.comboBox_4.setStyleSheet("font-size:29px;")
                self.comboBox_4.setObjectName("comboBox_4")
                self.comboBox_4.addItems(self.red_choice)
                self.comboBox_4.setMaxVisibleItems(self.comboBox_4.count())
                self.comboBox_4.setVisible(False)

                self.comboBox_5 = QtWidgets.QComboBox(self)
                self.comboBox_5.setGeometry(QtCore.QRect(200, 410-adjust_height, 150, 51))
                self.comboBox_5.setStyleSheet("font-size:29px;")
                self.comboBox_5.setObjectName("comboBox_5")
                self.comboBox_5.addItems(self.green_choice)
                self.comboBox_5.setVisible(False)

                # Melting date                                                                   
                self.label_2 = QtWidgets.QLabel(self)
                self.label_2.setGeometry(QtCore.QRect(20, 130-adjust_height, 271, 71))
                self.label_2.setStyleSheet("font-size : 28px; \n")
                self.label_2.setObjectName("label_2")

                self.dateEdit = QtWidgets.QDateEdit(self)
                self.dateEdit.setGeometry(QtCore.QRect(20, 190-adjust_height, 190, 51))
                self.dateEdit.setStyleSheet("font-size : 50px;")
                self.dateEdit.setDateTime(QtCore.QDateTime(QtCore.QDate(2023, 2, 1), QtCore.QTime(0, 0, 0)))
                self.dateEdit.setObjectName("dateEdit")
                
                # setting date now to the date edit
                d = QDate(now.year, now.month, now.day)
                self.dateEdit.setDate(d)
                self.dateEdit.setCalendarPopup(True)

                # Choosing shift
                self.label_3 = QtWidgets.QLabel(self)
                self.label_3.setGeometry(QtCore.QRect(268, 130-adjust_height, 131, 71))
                self.label_3.setStyleSheet("font-size : 30px;\n"
                                           "text-align: center;")
                self.label_3.setObjectName("label_3")

                list_shift=['1','2','3']
                self.comboBox_2 = QtWidgets.QComboBox(self, )
                self.comboBox_2.setGeometry(QtCore.QRect(225, 190-adjust_height, 160, 51))
                self.comboBox_2.setStyleSheet("font-size : 40px;")
                self.comboBox_2.setObjectName("comboBox_2")
                self.comboBox_2.addItems(list_shift)

                # pyhon date
                self.label_5 = QtWidgets.QLabel(self)
                self.label_5.setGeometry(QtCore.QRect(770, 90, 700, 51))
                self.label_5.setStyleSheet("font-size : 40px;")
                self.label_5.setObjectName("label_5")

                # Kg Unit
                self.label_6 = QtWidgets.QLabel(self)
                self.label_6.setGeometry(QtCore.QRect(900, 400-adjust_height, 251, 111))
                self.label_6.setStyleSheet("font-size : 75px;")
                self.label_6.setObjectName("label_6")

                # quantity based on weight scale
                self.lcdNumber = QtWidgets.QLCDNumber(self)
                self.lcdNumber.setGeometry(QtCore.QRect(397, 280, 501, 251))
                self.lcdNumber.setSmallDecimalPoint(True)  # display only one decimal point
                self.lcdNumber.setStyleSheet("color: white;\n"
                                                "padding: 15px 32px;\n"
                                                "font-weight : bold;\n"
                                                "text-align: right;\n"
                                                "font-size: 400px;")
                self.lcdNumber.setObjectName("lcdNumber")
                
                # Submit button
                self.pushButton_2 = QtWidgets.QPushButton(self, clicked = lambda: self.update())
                self.pushButton_2.setGeometry(QtCore.QRect(20, 480-adjust_height, 261, 91))
                self.pushButton_2.setStyleSheet("  background-color: #4CAF50; /* Green */\n"
                                                "  border: none;\n"
                                                "  font-weight : bold; \n"
                                                "  color: white;\n"
                                                "  padding: 15px 32px;\n"
                                                "  text-align: center;\n"
                                                "  font-size: 50px;")
                self.pushButton_2.setObjectName("pushButton_2")

                self.pushButton.setText(_translate("MainWindow", "Pako Smart tScale"))
                self.label.setText(_translate("MainWindow", "Pilih Material"))
                self.label_2.setText(_translate("MainWindow", "Tanggal Melting"))
                self.label_3.setText(_translate("MainWindow", "Shift"))
                self.label_4.setText(_translate("MainWindow", "Lini"))
                self.pushButton_2.setText(_translate("MainWindow", "Kirim"))
                #self.pushButton_3.setText(_translate("MainWindow", "Separator"))
                self.label_6.setText(_translate("MainWindow", "Kg"))
                self.label_7.setText(_translate("MainWindow", "Hijau"))
                self.label_8.setText(_translate("MainWindow", "Merah"))
                self.setWindowTitle(_translate("MainWindow", "Timbangan ZPF"))

                QtCore.QMetaObject.connectSlotsByName(self)

        def get_query(self, request):
                mycursor = self.mydb.cursor()
                query = request
                mycursor.execute(query)

                item = []
                for row in mycursor:
                    item.append(row) 
                return item
        
        def update_display(self):
                self.stop_event = threading.Event()
                self.thread = threading.Thread(target=self.display)
                self.thread.start()

        def display(self):
                _translate = QtCore.QCoreApplication.translate
                line = self.comboBox_3.currentText()
                if(line == "separator"):
                        if(self.separator_count < 1):
                                self.visible_choice(True)
                                self.comboBox.clear()
                                self.comboBox.addItems(self.list_separator)
                                self.separator_count += 1
                                self.melting_count = 0

                if(line[:7] == "melting"):  
                        if(self.melting_count < 1):
                                self.visible_choice(False)
                                self.comboBox.clear()
                                self.comboBox.addItems(self.list_items_melt)
                                self.separator_count = 0
                                self.melting_count += 1 

                # update time
                now = datetime.datetime.now()
                self.label_5.setText(_translate("MainWindow",now.strftime("%d-%m-%Y"), None))

                # update weight scale
                self.scale_value = get_scale_value()
                #self.random_index = random.randint(0,len(berat_timbangan)-1)
                #self.scale_value = berat_timbangan[self.random_index]
                self.lcdNumber.display(_translate("MainWindow", str(self.scale_value)))
                self.count = 0
                self.stop_event.set()
        
        def visible_choice(self, condition):
                        self.comboBox_4.setVisible(condition)
                        self.comboBox_5.setVisible(condition)
                        self.label_7.setVisible(condition)
                        self.label_8.setVisible(condition) 

        def update(self):
                _translate = QtCore.QCoreApplication.translate
                line = self.comboBox_3.currentText()
                # update weight scale
                if(self.convert_item_to_code[self.comboBox.currentText()] == "RA00000077"):
                        self.red_box = int(self.comboBox_4.currentText()) * weight_list[0]
                        self.green_box = int(self.comboBox_5.currentText()) * weight_list[1]
                        self.trolley = weight_list[2]
                        self.berat = float(get_scale_value())
                        #self.scale_value = float(berat_timbangan[self.random_index]) - self.red_box - self.green_box - self.trolley
                        self.berat_total =  self.berat - (self.red_box + self.green_box + self.trolley)
                        self.berat_total = format(self.berat_total, ".2f")
                        print("Berat timbangan separator : ", end="")
                        print(self.berat_total)
                        print("Berat Timbangan = {} \t Berat box merah = {} \t Berat box hijau = {} \t Berat Trolley = {}".format(self.berat, self.red_box, self.green_box, self.trolley))
                else:
                        self.berat_total = get_scale_value()
                        print("Berat timbangan melting : ", end="") 
                        print(self.berat_total)
                        
                message_information = "Berat = " + str(self.berat_total) + " Kg\n" + "Lini = " + self.comboBox_3.currentText() + "\nShift = " + str(self.comboBox_2.currentText()) + "\n\nBerhasil Diupload !!!"
                
                # prevent backdate
                day = self.dateEdit.date().toPyDate().day
                shift = str(self.comboBox_2.currentText()) 
                data = [str(self.dateEdit.date().toPyDate()), str(self.comboBox_2.currentText()), str(self.comboBox_3.currentText()), str(self.comboBox.currentText()), self.convert_item_to_code[self.comboBox.currentText()], str(self.scale_value)]
                self.update_data(data, message_information)

                #if(shift == '3'):
                #        if(day >= datetime.datetime.now().day-1 and day < datetime.datetime.now().day+1): 
                #                # update data
                #                self.update_data(data, message_information)
                #        else:
                #                self.fail_notification()

                #else:
                #        if(day is datetime.datetime.now().day):
                                # update data
                #                self.update_data(data, message_information)
                #        else:
                #                self.fail_notification()  
                
        def update_data(self, data, message):
                
                self.data = data
                self.message = message

                try:
                # update excel
                        self.update_excel(self.data)
                        
                        # Update DB
                        mycursor = self.mydb.cursor()
                        sql = """insert into transaction (DATETIME, TANGGAL_MELTING, SHIFT, LINI_MELTING, JENIS_MATERIAL, ITEM_CODE, BERAT) values ('%s','%s','%s','%s','%s','%s','%s')"""%(str(datetime.datetime.now()), self.data[0], self.data[1], self.data[2], self.data[3], self.data[4], self.data[5])
                        mycursor.execute(sql)
                        self.mydb.commit()
                        self.notification("Success", self.message)
                except self.db.errors.InterfaceError as err:
                        # Handle the connection error
                        print("Pesan Error upload : ", err)
                        self.notification("Gagal Upload !!", 'Silahkan menekan ulang tombol "Kirim"')
                        # wait for 5 second before attempting to reconnect 
                        time.sleep(5)
                        # reconnect to the server 
                        self.mydb =  self.db.connect()

                        # counting error 
                        self.count_error += 1
                        print("Error upload DB = ", end = '')
                        print(self.count_error)

        
        # ---------------------------------- Backup excel data -------------------------------
        # Every month
        def update_excel(self, data):
                year_now = str(datetime.datetime.now().year)
                month_now = str(datetime.datetime.now().month)
                date_now = month_now + "-" + year_now
                file_name = "backup_data/" + date_now + "_logger.xlsx"
                if os.path.exists(file_name):
                        wb = load_workbook(filename = file_name)
                        sheet = wb.active
                else:
                        wb = Workbook()
                        sheet = wb.active
                sheet['A1'] = "Datetime"
                sheet['B1'] = "Tanggal Melting"
                sheet['C1'] = 'Shift'
                sheet['D1'] = 'Lini_Melting'
                sheet['E1'] = 'Jenis_material'
                sheet['F1'] = 'Item Code'
                sheet['G1'] = 'Berat (Kg)'
                max = sheet.max_row
                
                # Data prepared before update
                data_update = str(datetime.datetime.now())[:19]
                rows = [data[0], data[1], data[2], data[3], data[4], data[5]]
                
                # Update data
                sheet.cell(row = 1 + max, column = 1).value = data_update
                for c, value in enumerate(rows, start = 1):
                        sheet.cell(row = 1 + max, column = c+1).value = value
                        
                #save the file
                wb.save(filename=file_name)
                wb.close()
        
        def notification(self, message, information):
                # notification
                qmsgBox = QMessageBox()
                qmsgBox.setStyleSheet('width = 400px')
                qmsgBox.information(self, message, information)

if __name__ == "__main__":
        app = QtWidgets.QApplication(sys.argv)

        # auto update weight timer 
        waktu = QtCore.QTimer()

        #instanciate class
        ui = meltingChip(waktu)
        
        # Create a stacked widget and add 2 main window
        stacked_widget = QStackedWidget()
        stacked_widget.addWidget(ui)
        
        # Set the size of the parent widget 
        parent_widget = QWidget()
        parent_widget.resize(1024, 600)
        parent_widget.setStyleSheet("background-color : #072227; color : white; font-size : 40px;")

        # set stacked widget as the main widget of the parent widget 
        layout = QVBoxLayout()
        layout.addWidget(stacked_widget)
        parent_widget.setLayout(layout)
        
        parent_widget.show()
        sys.exit(app.exec_())
